import requests

from openpyxl import load_workbook, Workbook

loadbook = load_workbook('test.xlsx')
ws = loadbook['Sheet1']
wb=Workbook()
sheet=wb.active


# for row in ws.iter_rows():
#   for cell in row:
    # dostęp do wartości komórek -> cell.internal_value
    # r = requests.get(cell.internal_value)
    # print(r.status_code)
    # to samo tylko do arraya jakiegos i potem wczytać
    

data=[('Id','Name','Marks'),
      (1,"ABC",50),
      (2,"CDE",100)]
# append all rows
for row in data:
    sheet.append(row)

wb.save("test.xlsx")